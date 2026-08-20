from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE  = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD  = Font(name="Arial", size=10, bold=True)
H1    = Font(name="Arial", size=14, bold=True)
H2    = Font(name="Arial", size=11, bold=True)
NOTE  = Font(name="Arial", size=9, italic=True, color="595959")
YEL   = PatternFill("solid", fgColor="FFFF00")
GREY  = PatternFill("solid", fgColor="F2F2F2")
thin  = Side(style="thin", color="BFBFBF")
BOX   = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR = '$#,##0.00;($#,##0.00);-'
CUR0= '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'
MULT= '0.00x'
NUM = '#,##0;(#,##0);-'

wb = Workbook()

# ============ SHEET 1: INPUTS ============
s = wb.active
s.title = "Inputs"
s.sheet_view.showGridLines = False
s["A1"] = "HIVOLT — Unit Economics Model"; s["A1"].font = H1
s["A2"] = "Fill every YELLOW cell from your Tapstitch account. Blue cells are levers you can change. Black cells are formulas — do not edit."
s["A2"].font = NOTE

s["A4"] = "LEGEND"; s["A4"].font = H2
s["A5"] = "Yellow fill"; s["A5"].font = BOLD; s["A5"].fill = YEL
s["B5"] = "You must supply this from Tapstitch (blank cost, print fee, shipping, duty)"; s["B5"].font = NOTE
s["A6"] = "Blue text"; s["A6"].font = BLUE
s["B6"] = "Assumption / lever — change to run scenarios"; s["B6"].font = NOTE
s["A7"] = "Black text"; s["A7"].font = BLACK
s["B7"] = "Calculated — leave alone"; s["B7"].font = NOTE

# --- Tapstitch per-SKU costs ---
s["A9"] = "1. TAPSTITCH COSTS PER PRODUCT"; s["A9"].font = H2
s["A10"] = "Pull from Tapstitch > catalog (select US fulfilment region) and their Price Checker tool."
s["A10"].font = NOTE

hdr = ["Product","Retail $","# Blanks","Blank cost ea $","# Print areas","Print fee/area $",
       "Ship 1st item $","Ship addl item $","Duty + svc fee $","US fulfil?"]
for i,h in enumerate(hdr, start=1):
    c = s.cell(row=12, column=i, value=h)
    c.font = BOLD; c.fill = GREY; c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Real HIVOLT products + real retail prices. Costs blank for user.
prods = [
    ("Voltcore 2-Piece Set (bra + leggings)", 79.00, 2),
    ("Women's High-Waisted Flare Leggings",   54.00, 1),
    ("Women's Twist Front V-Neck Sports Bra", 38.00, 1),
    ("Men's High-Stretch Gym T-Shirt",        34.00, 1),
    ("Men's Tapered Leg Drawstring Joggers",  69.00, 1),
    ("Women's High Rise Yoga Shorts",         42.00, 1),
    ("Two-Tone Fleeced Varsity Jacket",       69.00, 1),
]
r = 13
for name, price, blanks in prods:
    s.cell(row=r, column=1, value=name).font = BLACK
    s.cell(row=r, column=2, value=price).font = BLUE
    s.cell(row=r, column=2).number_format = CUR
    s.cell(row=r, column=3, value=blanks).font = BLACK
    for col in (4,5,6,7,8,9,10):
        c = s.cell(row=r, column=col)
        c.fill = YEL; c.font = BLACK
        if col in (4,6,7,8,9): c.number_format = CUR
    for col in range(1,11):
        s.cell(row=r, column=col).border = BOX
    r += 1

EX = r  # example row
s.cell(row=EX, column=1, value="EXAMPLE (delete) — Leggings WB0031").font = NOTE
for col,val in [(2,54.00),(3,1),(4,12.50),(5,1),(6,2.99),(7,4.44),(8,0.50),(9,2.00)]:
    c = s.cell(row=EX, column=col, value=val); c.font = NOTE
    if col in (2,4,6,7,8,9): c.number_format = CUR
s.cell(row=EX, column=10, value="No").font = NOTE
s.cell(row=EX+1, column=1, value="Example values are illustrative only — NOT real Tapstitch prices. Research could not obtain published activewear costs.").font = NOTE

# --- Store assumptions ---
A = EX + 3
s.cell(row=A, column=1, value="2. STORE ASSUMPTIONS").font = H2
rows = [
    ("Payment fee %",            0.029, PCT, "Shopify Payments US standard ~2.9%"),
    ("Payment fee fixed $",       0.30, CUR, "Per transaction"),
    ("Return rate %",             0.26, PCT, "US apparel avg ~26%. Tapstitch reviews cite sizing runs small — consider 0.30+"),
    ("Recovery on returns %",     0.00, PCT, "POD returns usually not resellable. 0% = total loss of goods."),
    ("Discount uptake %",         0.50, PCT, "Share of orders using VOLT20 / WELCOME10"),
    ("Avg discount %",            0.20, PCT, "VOLT20 = 20%"),
]
s.cell(row=A+1, column=1, value="Assumption").font = BOLD
s.cell(row=A+1, column=2, value="Value").font = BOLD
s.cell(row=A+1, column=3, value="Note").font = BOLD
for i,(lbl,val,fmt,note) in enumerate(rows, start=A+2):
    s.cell(row=i, column=1, value=lbl).font = BLACK
    c = s.cell(row=i, column=2, value=val); c.font = BLUE; c.number_format = fmt
    s.cell(row=i, column=3, value=note).font = NOTE
ASSUM = A+2  # first assumption row

# --- Acquisition assumptions ---
B = A + 9
s.cell(row=B, column=1, value="3. ACQUISITION ASSUMPTIONS").font = H2
arows = [
    ("Meta CPC $",            0.55, CUR, "Sourced: apparel CPC $0.45-0.58 (Triple Whale / Lebesgue, CY2025)"),
    ("Conversion rate %",     0.010, PCT, "Sourced: new-brand cold paid social 0.7-1.2%. 1.0% = midpoint"),
    ("TikTok referral fee %", 0.06,  PCT, "Sourced: 6% US flat; some sources cite 8% apparel. Verify in Seller Center"),
    ("TikTok affiliate comm %",0.16, PCT, "Sourced: fashion 15-18%; US avg 13.02%"),
]
s.cell(row=B+1, column=1, value="Assumption").font = BOLD
s.cell(row=B+1, column=2, value="Value").font = BOLD
s.cell(row=B+1, column=3, value="Source").font = BOLD
for i,(lbl,val,fmt,note) in enumerate(arows, start=B+2):
    s.cell(row=i, column=1, value=lbl).font = BLACK
    c = s.cell(row=i, column=2, value=val); c.font = BLUE; c.number_format = fmt
    s.cell(row=i, column=3, value=note).font = NOTE
ACQ = B+2

s.column_dimensions["A"].width = 38
for col in "BCDEFGHIJ": s.column_dimensions[col].width = 13
s.column_dimensions["C"].width = 60 if False else 13

# store row refs for other sheets
FIRST, LAST = 13, 13+len(prods)-1
PAY_PCT, PAY_FIX, RET_RATE, RECOV, DISC_UP, DISC_PCT = (ASSUM, ASSUM+1, ASSUM+2, ASSUM+3, ASSUM+4, ASSUM+5)
CPC, CVR, TT_REF, TT_AFF = (ACQ, ACQ+1, ACQ+2, ACQ+3)

# ============ SHEET 2: UNIT ECONOMICS ============
u = wb.create_sheet("Unit Economics")
u.sheet_view.showGridLines = False
u["A1"] = "Unit Economics by Product"; u["A1"].font = H1
u["A2"] = "Everything here is formula-driven from Inputs. Fill the yellow cells there and this populates."
u["A2"].font = NOTE

uh = ["Product","Net revenue","Landed COGS","Shipping","Payment fees","Returns cost",
      "CM2 ($)","CM2 (%)","Breakeven ROAS","Allowable CAC","Actual CAC @ CPC/CVR","Profit / order"]
for i,h in enumerate(uh, start=1):
    c = u.cell(row=4, column=i, value=h)
    c.font = BOLD; c.fill = GREY; c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

for n in range(len(prods)):
    src = FIRST + n
    row = 5 + n
    u.cell(row=row, column=1, value=f"=Inputs!A{src}").font = BLACK
    # Net revenue after blended discount
    u.cell(row=row, column=2, value=f"=Inputs!B{src}*(1-Inputs!$B${DISC_UP}*Inputs!$B${DISC_PCT})")
    # Landed COGS = blanks*blankcost + printareas*printfee + duty/svc
    u.cell(row=row, column=3, value=f"=Inputs!C{src}*Inputs!D{src}+Inputs!E{src}*Inputs!F{src}+Inputs!I{src}")
    # Shipping = first + (blanks-1)*addl
    u.cell(row=row, column=4, value=f"=Inputs!G{src}+MAX(0,Inputs!C{src}-1)*Inputs!H{src}")
    # Payment fees on net revenue
    u.cell(row=row, column=5, value=f"=B{row}*Inputs!$B${PAY_PCT}+Inputs!$B${PAY_FIX}")
    # Returns: lose goods + outbound shipping, refund revenue
    u.cell(row=row, column=6, value=f"=Inputs!$B${RET_RATE}*(C{row}+D{row})*(1-Inputs!$B${RECOV})")
    # CM2
    u.cell(row=row, column=7, value=f"=B{row}-C{row}-D{row}-E{row}-F{row}")
    u.cell(row=row, column=8, value=f"=IFERROR(G{row}/B{row},0)")
    u.cell(row=row, column=9, value=f"=IFERROR(1/H{row},0)")
    u.cell(row=row, column=10, value=f"=G{row}")
    u.cell(row=row, column=11, value=f"=IFERROR(Inputs!$B${CPC}/Inputs!$B${CVR},0)")
    u.cell(row=row, column=12, value=f"=G{row}-K{row}")
    for col,fmt in [(2,CUR),(3,CUR),(4,CUR),(5,CUR),(6,CUR),(7,CUR),(8,PCT),(9,MULT),(10,CUR),(11,CUR),(12,CUR)]:
        u.cell(row=row, column=col).number_format = fmt
        u.cell(row=row, column=col).font = BLACK
    for col in range(1,13):
        u.cell(row=row, column=col).border = BOX

VER = 5+len(prods)+2
u.cell(row=VER, column=1, value="HOW TO READ THIS").font = H2
notes = [
 ("CM2 (%)", "Contribution margin before ad spend. Sourced threshold: below 30% \"paid acquisition gets very hard\"."),
 ("Breakeven ROAS", "= 1 / CM2%. The return on ad spend you must beat just to not lose money."),
 ("Allowable CAC", "The most you can pay to acquire one order and break even. Equals CM2."),
 ("Actual CAC", "= CPC / conversion rate. What Meta will actually charge you at the assumed inputs."),
 ("Profit / order", "Allowable CAC minus Actual CAC. NEGATIVE means you lose money on every order sold."),
]
for i,(a,b) in enumerate(notes, start=VER+1):
    u.cell(row=i, column=1, value=a).font = BOLD
    u.cell(row=i, column=2, value=b).font = NOTE

u.column_dimensions["A"].width = 36
for col in "BCDEFGHIJKL": u.column_dimensions[col].width = 14

# ============ SHEET 3: REVENUE TARGET ============
t = wb.create_sheet("Revenue Target")
t.sheet_view.showGridLines = False
t["A1"] = "Revenue Target Model"; t["A1"].font = H1
t["A2"] = "What $250k/month actually requires, using the unit economics from the previous sheet."
t["A2"].font = NOTE

t["A4"] = "Target monthly revenue"; t["A4"].font = BOLD
t["B4"] = 250000; t["B4"].font = BLUE; t["B4"].number_format = CUR0
t["C4"] = "Change this to model any target"; t["C4"].font = NOTE

t["A5"] = "Blended AOV"; t["A5"].font = BOLD
t["B5"] = f"=IFERROR(AVERAGE('Unit Economics'!B5:B{4+len(prods)}),0)"
t["B5"].number_format = CUR; t["B5"].font = BLACK
t["C5"] = "Average net revenue across catalogue. Raising this is the highest-leverage lever."; t["C5"].font = NOTE

t["A6"] = "Blended CM2 per order"; t["A6"].font = BOLD
t["B6"] = f"=IFERROR(AVERAGE('Unit Economics'!G5:G{4+len(prods)}),0)"
t["B6"].number_format = CUR; t["B6"].font = BLACK

t["A7"] = "Blended CM2 %"; t["A7"].font = BOLD
t["B7"] = "=IFERROR(B6/B5,0)"; t["B7"].number_format = PCT; t["B7"].font = BLACK

t["A9"] = "VOLUME REQUIRED"; t["A9"].font = H2
vol = [
 ("Orders per month",    "=IFERROR(B4/B5,0)", NUM),
 ("Orders per day",      "=IFERROR(B10/30,0)", NUM),
 ("Sessions/mo @ 0.7% CVR","=IFERROR(B10/0.007,0)", NUM),
 ("Sessions/mo @ 1.0% CVR","=IFERROR(B10/0.010,0)", NUM),
 ("Sessions/mo @ 2.0% CVR","=IFERROR(B10/0.020,0)", NUM),
]
for i,(lbl,f,fmt) in enumerate(vol, start=10):
    t.cell(row=i, column=1, value=lbl).font = BLACK
    c = t.cell(row=i, column=2, value=f); c.number_format = fmt; c.font = BLACK
t["C12"] = "Sourced: new-brand cold paid social converts 0.7-1.2%"; t["C12"].font = NOTE

t["A16"] = "PAID ACQUISITION OUTCOME"; t["A16"].font = H2
paid = [
 ("Actual CAC (CPC / CVR)",  f"=IFERROR(Inputs!B{CPC}/Inputs!B{CVR},0)", CUR),
 ("Ad spend to hit target",  "=IFERROR(B10*B17,0)", CUR0),
 ("Contribution generated",  "=IFERROR(B10*B6,0)", CUR0),
 ("NET PROFIT / (LOSS)",     "=B19-B18", CUR0),
 ("Achieved ROAS",           "=IFERROR(B4/B18,0)", MULT),
 ("Breakeven ROAS required", "=IFERROR(1/B7,0)", MULT),
]
for i,(lbl,f,fmt) in enumerate(paid, start=17):
    t.cell(row=i, column=1, value=lbl).font = BOLD if "NET" in lbl else BLACK
    c = t.cell(row=i, column=2, value=f); c.number_format = fmt
    c.font = BOLD if "NET" in lbl else BLACK

t["A24"] = "TIKTOK SHOP ALTERNATIVE"; t["A24"].font = H2
tt = [
 ("Platform take (referral + affiliate)", f"=Inputs!B{TT_REF}+Inputs!B{TT_AFF}", PCT),
 ("Cost per order on TikTok",             f"=B5*(Inputs!B{TT_REF}+Inputs!B{TT_AFF})", CUR),
 ("vs Meta actual CAC",                   "=B17-B26", CUR),
]
for i,(lbl,f,fmt) in enumerate(tt, start=25):
    t.cell(row=i, column=1, value=lbl).font = BLACK
    c = t.cell(row=i, column=2, value=f); c.number_format = fmt; c.font = BLACK
t["C27"] = "Positive = TikTok Shop is cheaper per order than Meta at these assumptions"; t["C27"].font = NOTE
t["A29"] = "Reality check: average active US TikTok Shop seller does $3,750/month GMV. Top 10% of sellers drive ~48% of all GMV."
t["A29"].font = NOTE
t["A30"] = "$250k/month is top-decile TikTok Shop performance and is a content-volume business (10+ creator videos/week at scale)."
t["A30"].font = NOTE

t.column_dimensions["A"].width = 34
t.column_dimensions["B"].width = 16
t.column_dimensions["C"].width = 62

# ============ SHEET 4: SOURCES ============
v = wb.create_sheet("Sources")
v.sheet_view.showGridLines = False
v["A1"] = "Sources & Confidence"; v["A1"].font = H1
v["A2"] = "Both research passes were run via web search only — direct page fetching was blocked by the network proxy. Figures are attributed but page methodology could not be verified."
v["A2"].font = NOTE
sh = ["Figure","Value","Confidence","Source"]
for i,h in enumerate(sh, start=1):
    c = v.cell(row=4, column=i, value=h); c.font = BOLD; c.fill = GREY; c.border = BOX
src = [
 ("Tapstitch activewear blank cost","NOT FOUND","n/a","Prices render client-side; not in search index. Must pull from your account."),
 ("Tapstitch US shipping","$4.44 first item + $0.50 addl","HIGH — first party","tapstitch.com/help-center/faq/detail/shipping-cost"),
 ("Tapstitch print fee","$2.99/area OR ~$5.00 one-side","LOW — sources conflict","michaelessek.com ; printsgram.com"),
 ("Tapstitch US tariff surcharge","20% duty on blank + $0.80-2.00/pc","HIGH — first party","support.tapstitch.com — U.S. Tariff Adjustment notice"),
 ("Tapstitch activewear launch","March 2026 — US stocking partial","HIGH — first party","The Monthly Stitch, March 2026"),
 ("Apparel Meta CPC","$0.45 - $0.58","MEDIUM","Triple Whale ; Lebesgue"),
 ("New-brand cold paid social CVR","0.7% - 1.2%","MEDIUM","Eightx conversion-rate-by-traffic-source 2026"),
 ("Fashion Meta ROAS median","2.18x (top 10% = 6.0x)","MEDIUM","AdAmigo fashion ROAS benchmarks 2026"),
 ("US apparel return rate","20% - 40% (avg ~26%)","MEDIUM","NRF / Happy Returns ; Richpanel"),
 ("CM2 threshold for paid to work","30% minimum","MEDIUM","ask-luca.com contribution margin"),
 ("TikTok Shop referral fee","6% US flat (some cite 8% apparel)","LOW — conflicting","Fit Small Business ; Darkroom ; Dashboardly"),
 ("TikTok fashion affiliate commission","15% - 18% (US avg 13.02%)","MEDIUM","Shortform Nation 2026 data"),
 ("Avg active TikTok Shop seller GMV","$3,750/month","MEDIUM","Grey Journal 2026"),
]
for i,row in enumerate(src, start=5):
    for j,val in enumerate(row, start=1):
        c = v.cell(row=i, column=j, value=val)
        c.font = BLACK; c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
v.column_dimensions["A"].width = 34
v.column_dimensions["B"].width = 30
v.column_dimensions["C"].width = 20
v.column_dimensions["D"].width = 52

wb.save("HIVOLT-unit-economics.xlsx")
print("built")
